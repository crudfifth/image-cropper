import csv
from .models import JCreditApplication
import datetime
import openpyxl
import os
from copy import copy, deepcopy

from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent.parent
DIR_PATH_EXCEL_TEMPLATE = os.path.join(BASE_DIR,'ihiapp/excel_template')
DIR_PATH_EXCEL_OUTPUT = os.path.join(BASE_DIR,'ihiapp/excel_output')
class JCreditRowWriter:
    def __init__(self, ws , row , jcredit_application):
        self.ws = ws
        self.row = row
        self.jcredit_application = jcredit_application
        self.col = 1
    
    # １列目から順次セルに値を書き込む。呼び出すたびに列が進む。
    def write_value(self, value):
        if value is True:
            value = "○"
        elif value is False:
            value = "×"
        self.ws.cell(row=self.row, column=self.col, value=value)
        self.col += 1
    # 読み取りcolを先に進めるだけ
    def seek_col(self):
        self.col += 1

# JCreditApplicationのレコードをもとに、Excelファイルを生成する。
def generate_excel():
    # タイムスタンプ付きのファイル名を生成する
    timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    source_filepath = os.path.join( DIR_PATH_EXCEL_TEMPLATE, 'EN-S-001_sakugen_list_boiler_jigyousyo.xlsx' )
    destination_filepath = os.path.join( DIR_PATH_EXCEL_OUTPUT, 'jcredit_output_%s.xlsx'%(timestamp) )
    # excel_templateフォルダからテンプレートファイルを読み込む
    wb = openpyxl.load_workbook(source_filepath)
    # テンプレートファイルから会員情報シートを取得する
    ws = wb['会員情報']
    # 12行目からデータが始まるため、12行目からデータを取得する
    set_row = 12
    current_index = 1
    # JCreditApplicationのレコードを全て取得する
    for jcredit_application in JCreditApplication.objects.all():
        # 現在の行にjcredit_applicationのデータを書き込む
        RowWriter = JCreditRowWriter(ws, set_row, jcredit_application)
        # 頭から一個一個書き込む
        RowWriter.write_value(None)
        RowWriter.write_value(None)
        RowWriter.write_value(current_index)
        RowWriter.seek_col()
        RowWriter.write_value(jcredit_application.application_date)
        RowWriter.write_value(jcredit_application.member_name )
        RowWriter.write_value(jcredit_application.member_postal_code)
        RowWriter.write_value(jcredit_application.member_address)
        RowWriter.write_value(jcredit_application.member_phone_number)
        RowWriter.write_value(jcredit_application.installation_postal_code)
        RowWriter.write_value(jcredit_application.installation_address)
        RowWriter.write_value(jcredit_application.keidanren_carbon_neutral_participation)
        RowWriter.write_value(jcredit_application.energy_conservation_law_specified_business_number)
        RowWriter.write_value(jcredit_application.global_warming_countermeasures_specified_emitter_code)
        RowWriter.write_value(jcredit_application.base_manufacturer_name)
        RowWriter.write_value(jcredit_application.base_model)
        RowWriter.write_value(jcredit_application.base_quantity)
        RowWriter.write_value(jcredit_application.base_output)
        RowWriter.write_value(jcredit_application.base_unit_heat_gen_of_fossil_fuel_in_base_boiler)
        RowWriter.write_value(jcredit_application.base_efficiency_percentage)
        RowWriter.write_value(jcredit_application.base_efficiency_of_standard_equipment_percentage)
        RowWriter.write_value(jcredit_application.base_fuel_type)
        RowWriter.write_value(jcredit_application.base_legal_service_life_years)
        RowWriter.write_value(jcredit_application.base_installation_date)
        RowWriter.write_value(jcredit_application.base_operation_end_date)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.base_years_of_operation)
        RowWriter.write_value(jcredit_application.updated_manufacturer_name)
        RowWriter.write_value(jcredit_application.updated_model)
        RowWriter.write_value(jcredit_application.updated_quantity)
        RowWriter.write_value(jcredit_application.updated_unique_number)
        RowWriter.write_value(jcredit_application.updated_output)
        RowWriter.write_value(jcredit_application.updated_unit_heat_gen_of_fossil_fuel_in_base_boiler)
        RowWriter.write_value(jcredit_application.updated_efficiency_percentage)
        RowWriter.write_value(jcredit_application.updated_fuel_type)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.eval_investment_recovery_years)
        RowWriter.write_value(jcredit_application.eval_total_investment_amount)
        RowWriter.write_value(jcredit_application.eval_subsidy_amount)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.eval_net_investment_amount)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.eval_running_cost_before_implementation)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.eval_running_cost_after_implementation)
        RowWriter.write_value(jcredit_application.eval_baseline_fuel_unit_price_per_month)
        RowWriter.write_value(jcredit_application.eval_post_implementation_fuel_unit_price_per_month)
        RowWriter.write_value(jcredit_application.eval_existing_maintenance_cost_per_year)
        RowWriter.write_value(jcredit_application.eval_post_implementation_maintenance_cost_per_year)
        RowWriter.write_value(jcredit_application.eval_documentation)
        RowWriter.write_value(jcredit_application.eval_subsidy_name)
        RowWriter.write_value(jcredit_application.eval_granting_organization_of_subsidy)
        RowWriter.write_value(jcredit_application.eval_domestically_implemented_in_japan)
        RowWriter.write_value(jcredit_application.operating_start_date)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.certify_start_date)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.certify_end_date)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.requirement_implemented_in_japan)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.requirement_project_execution_date)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.requirement_certification_start_date)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.requirement_equipment_efficiency_higher_than_standard)
        RowWriter.seek_col() # RowWriter.write_value(jcredit_application.requirement_additivity)
        RowWriter.write_value(jcredit_application.unit_heat_gen_of_fossil_fuels_in_base_boiler)
        RowWriter.write_value(jcredit_application.co2_per_unit_heat_ge_fuels_in_base_boiler)
        RowWriter.write_value(jcredit_application.unit_heat_gen_of_fuel_used_in_boiler_after_project_exec)
        RowWriter.write_value(jcredit_application.co2_per_unit_heat_ge_fuels_used_in_boiler_after_exec)
        RowWriter.write_value(jcredit_application.monitoring_period_in_months)
        RowWriter.write_value(jcredit_application.monitoring_measurement_value_of_fuel_consumption)
        RowWriter.write_value(jcredit_application.monitoring_classification)
        RowWriter.write_value(jcredit_application.monitoring_measurement_error_rate_in_c)
        RowWriter.write_value(jcredit_application.monitoring_final_value)
        current_index += 1
        set_row += 1
        if current_index > 50:
            ws.insert_rows(set_row)
            # １つ上の行の値とスタイルを全てコピーする
            for col in range(1, 100):
                ws.cell(row=set_row, column=col).value = ws.cell(row=set_row-1, column=col).value
                ws.cell(row=set_row, column=col).font = ws.cell(row=set_row-1, column=col).font._StyleProxy__target
                ws.cell(row=set_row, column=col).border = copy(ws.cell(row=set_row-1, column=col).border)
                ws.cell(row=set_row, column=col).fill = ws.cell(row=set_row-1, column=col).fill._StyleProxy__target
                ws.cell(row=set_row, column=col).number_format = ws.cell(row=set_row-1, column=col).number_format
                ws.cell(row=set_row, column=col).protection = ws.cell(row=set_row-1, column=col).protection._StyleProxy__target
                ws.cell(row=set_row, column=col).alignment = ws.cell(row=set_row-1, column=col).alignment._StyleProxy__target
    # 最後の行は消す
    ws.delete_rows(set_row)
    
    # 全件書き込み終わったら、ファイルを保存する
    wb.save(filename = destination_filepath)
    return destination_filepath


    

